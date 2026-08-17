from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.cell import coordinate_from_string


def _values(ws, row: int) -> list[Any]:
    return [ws.cell(row, column).value for column in range(1, ws.max_column + 1)]


def _looks_like_table_header(values: list[Any]) -> bool:
    non_empty = [value for value in values if value not in (None, "")]
    return len(non_empty) >= 2 and all(isinstance(value, str) for value in non_empty)


def _find_header_row(ws, *, start: int = 1, stop: int = 15) -> int | None:
    for row in range(max(1, start), min(ws.max_row, stop) + 1):
        if _looks_like_table_header(_values(ws, row)):
            return row
    return None


def _freeze_row(value: object | None) -> int | None:
    if value is None:
        return None
    coordinate = getattr(value, "coordinate", None) or str(value)
    try:
        _, row = coordinate_from_string(coordinate)
    except (TypeError, ValueError):
        return None
    return row


def finalize_business_workbook(content: bytes) -> bytes:
    """Final readability pass for every Excel file exposed by the Mini App.

    The report builders already own the business content and ERA styling. This
    pass only corrects spreadsheet ergonomics: title rows must never become
    filter headers, filters should start on a real table header, freeze panes
    must not hide that header, and widths stay bounded for mobile/desktop use.
    No data is added and no hidden/raw psychology fields are introduced.
    """

    wb = load_workbook(BytesIO(content))
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

        if ws.auto_filter.ref:
            min_col, min_row, max_col, max_row = range_boundaries(ws.auto_filter.ref)
            if not _looks_like_table_header(_values(ws, min_row)):
                header_row = _find_header_row(ws, start=min_row + 1)
                if header_row is not None and header_row < ws.max_row:
                    ws.auto_filter.ref = (
                        f"{get_column_letter(min_col)}{header_row}:"
                        f"{get_column_letter(max_col)}{max_row}"
                    )
                    frozen_row = _freeze_row(ws.freeze_panes)
                    if frozen_row is None or frozen_row <= header_row:
                        ws.freeze_panes = f"A{header_row + 1}"

        for column in range(1, ws.max_column + 1):
            letter = get_column_letter(column)
            current = ws.column_dimensions[letter].width
            if current is None:
                continue
            ws.column_dimensions[letter].width = min(max(current, 10), 58)

        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
