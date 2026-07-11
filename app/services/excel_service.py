from __future__ import annotations

from collections import Counter
from collections.abc import Iterable
from datetime import date, datetime, time
from io import BytesIO
from typing import Any


def _plain(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return value


def _value(item: Any, key: str, default: Any = "") -> Any:
    if isinstance(item, dict):
        return item.get(key, default)
    return getattr(item, key, default)


def _label(value: Any, labels: dict[Any, str]) -> str:
    return labels.get(value, labels.get(str(value), str(value or "—")))


def _safe_percent(part: int, total: int) -> float:
    return round(part * 100 / total, 1) if total else 0.0


def _age_group(age: int | None) -> str:
    if age is None:
        return "Не указано"
    if age < 14:
        return "До 14"
    if age <= 17:
        return "14–17"
    if age <= 21:
        return "18–21"
    if age <= 25:
        return "22–25"
    return "26+"


def build_analytics_workbook(
    users,
    events,
    projects,
    point_totals: dict[int, int],
    *,
    department_stats: Iterable[Any] = (),
    direction_stats: Iterable[Any] = (),
    goals: Iterable[Any] = (),
    contacts: Iterable[Any] = (),
    sections: set[str] | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    from app.utils.constants import (
        APPLICATION_STATUS_LABELS,
        EVENT_STATUS_LABELS,
        PROJECT_STATUS_LABELS,
        ROLE_LABELS,
        STATUS_LABELS,
    )

    sections = sections or {
        "summary", "funnel", "demography", "activity", "users", "departments",
        "directions", "events", "projects", "goals", "contacts",
    }
    users = list(users)
    events = list(events)
    projects = list(projects)
    department_stats = list(department_stats)
    direction_stats = list(direction_stats)
    goals = list(goals)
    contacts = list(contacts)

    wb = Workbook()
    wb.remove(wb.active)

    purple = "6F2DBD"
    red = "E63946"
    green = "2A9D8F"
    amber = "F4A261"
    light = "F6F0FA"
    border = Border(bottom=Side(style="thin", color="DDD4E0"))
    table_index = 0

    def empty_row(width: int) -> list[str]:
        return ["Данных пока нет"] + [""] * max(width - 1, 0)

    def style_sheet(ws, freeze: str = "A2") -> None:
        nonlocal table_index
        ws.freeze_panes = freeze
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=purple)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            width = max(len(str(ws.cell(row, col).value or "")) for row in range(1, ws.max_row + 1))
            ws.column_dimensions[letter].width = min(max(width + 3, 12), 48)
        if ws.max_row > 1 and ws.max_column > 0:
            table_index += 1
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            table = Table(displayName=f"tbl_{table_index}", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            ws.add_table(table)

    def append_sheet(title: str, headers: list[str], rows: list[list[Any]]):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for row in rows or [empty_row(len(headers))]:
            ws.append([_plain(value) for value in row])
        style_sheet(ws)
        return ws

    approved_users = [u for u in users if str(getattr(u, "application_status", "")) == "approved"]
    pending_users = [u for u in users if str(getattr(u, "application_status", "")) == "pending"]
    active_users = [u for u in users if point_totals.get(u.id, 0) > 0]
    total_points_value = sum(point_totals.values())
    avg_points = round(total_points_value / len(users), 1) if users else 0
    median_points = 0
    if users:
        sorted_points = sorted(point_totals.get(u.id, 0) for u in users)
        mid = len(sorted_points) // 2
        median_points = sorted_points[mid] if len(sorted_points) % 2 else round((sorted_points[mid - 1] + sorted_points[mid]) / 2, 1)

    app_counts = Counter(_label(getattr(u, "application_status", ""), APPLICATION_STATUS_LABELS) for u in users)
    role_counts = Counter(_label(getattr(u, "role", ""), ROLE_LABELS) for u in users)
    status_counts = Counter(_label(getattr(u, "participation_status", ""), STATUS_LABELS) for u in users)
    city_counts = Counter((getattr(u, "city", None) or "Не указано").strip() for u in users)
    age_counts = Counter(_age_group(getattr(u, "age", None)) for u in users)
    event_status_counts = Counter(_label(getattr(e, "status", ""), EVENT_STATUS_LABELS) for e in events)
    project_status_counts = Counter(_label(getattr(p, "status", ""), PROJECT_STATUS_LABELS) for p in projects)

    if "summary" in sections:
        ws = wb.create_sheet("Сводка")
        ws.sheet_view.showGridLines = False
        ws["A1"] = "АНАЛИТИКА ЭРА"
        ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=red)
        ws.merge_cells("A1:D1")
        ws["A2"] = f"Сформировано: {datetime.now():%d.%m.%Y %H:%M}"
        ws.merge_cells("A2:D2")
        ws.append([])
        ws.append(["Ключевой показатель", "Значение", "Доля / динамика", "Управленческий смысл"])
        kpis = [
            ("Участники в базе", len(users), "100%", "Размер текущего сообщества"),
            ("Одобрены", len(approved_users), f"{_safe_percent(len(approved_users), len(users))}%", "Прошли вход в систему ЭРА"),
            ("Новые заявки", len(pending_users), f"{_safe_percent(len(pending_users), len(users))}%", "Требуют решения команды"),
            ("Участники с баллами", len(active_users), f"{_safe_percent(len(active_users), len(users))}%", "Есть подтверждённая активность"),
            ("Мероприятия", len(events), "—", "Объём событийной работы"),
            ("Проекты", len(projects), "—", "Инициативность сообщества"),
            ("Баллы всего", total_points_value, "—", "Суммарный подтверждённый вклад"),
            ("Средний баланс", avg_points, "—", "Средняя активность на участника"),
            ("Медианный баланс", median_points, "—", "Типичный уровень активности без влияния лидеров"),
            ("Организации", len(contacts), "—", "Размер внешней базы контактов"),
            ("Цели месяца", len(goals), "—", "Количество управленческих целей"),
        ]
        for row in kpis:
            ws.append(row)
        for cell in ws[4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=purple)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 48
        ws.freeze_panes = "A5"

        start = ws.max_row + 3
        ws.cell(start, 1, "Статусы заявок")
        ws.cell(start, 2, "Количество")
        for index, (name, count) in enumerate(app_counts.items(), start + 1):
            ws.cell(index, 1, name)
            ws.cell(index, 2, count)
        if app_counts:
            chart = PieChart()
            chart.title = "Структура заявок"
            chart.add_data(Reference(ws, min_col=2, min_row=start, max_row=start + len(app_counts)), titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(app_counts)))
            chart.height = 7
            chart.width = 10
            ws.add_chart(chart, "F4")

        top_start = ws.max_row + 3
        ws.cell(top_start, 1, "Топ-10 участников")
        ws.cell(top_start, 2, "Баллы")
        leaders = sorted(users, key=lambda u: point_totals.get(u.id, 0), reverse=True)[:10]
        for index, member in enumerate(leaders, top_start + 1):
            ws.cell(index, 1, f"{member.first_name} {member.last_name or ''}".strip())
            ws.cell(index, 2, point_totals.get(member.id, 0))
        if leaders:
            chart = BarChart()
            chart.title = "Топ-10 по активности"
            chart.y_axis.title = "Баллы"
            chart.add_data(Reference(ws, min_col=2, min_row=top_start, max_row=top_start + len(leaders)), titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=1, min_row=top_start + 1, max_row=top_start + len(leaders)))
            chart.height = 8
            chart.width = 14
            ws.add_chart(chart, "F19")

    if "funnel" in sections:
        funnel_rows = [
            ["Все участники", len(users), 100.0, "База сообщества"],
            ["Одобрены", len(approved_users), _safe_percent(len(approved_users), len(users)), "Прошли вход"],
            ["Выбрали департамент", sum(1 for u in users if getattr(u, "departments", None)), _safe_percent(sum(1 for u in users if getattr(u, "departments", None)), len(users)), "Определили направление участия"],
            ["Выбрали направление", sum(1 for u in users if getattr(u, "directions", None)), _safe_percent(sum(1 for u in users if getattr(u, "directions", None)), len(users)), "Есть предметный интерес"],
            ["Получили баллы", len(active_users), _safe_percent(len(active_users), len(users)), "Совершили подтверждённое действие"],
        ]
        ws = append_sheet("Воронка роста", ["Этап", "Количество", "Конверсия, %", "Что показывает"], funnel_rows)
        chart = BarChart()
        chart.title = "Воронка вовлечения"
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=ws.max_row), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        chart.height = 8
        chart.width = 14
        ws.add_chart(chart, "F2")

    if "demography" in sections:
        rows = []
        for category, counter in (("Возраст", age_counts), ("Город", city_counts), ("Роль", role_counts), ("Статус участия", status_counts)):
            total = sum(counter.values())
            rows.extend([[category, name, count, _safe_percent(count, total)] for name, count in counter.most_common()])
        append_sheet("Срезы аудитории", ["Срез", "Группа", "Количество", "Доля, %"], rows)

    if "activity" in sections:
        ranked = sorted(users, key=lambda u: point_totals.get(u.id, 0), reverse=True)
        rows = []
        for place, user in enumerate(ranked, 1):
            points = point_totals.get(user.id, 0)
            level = "Лидер активности" if points >= 500 else "Активный" if points >= 150 else "Вовлечённый" if points > 0 else "Старт"
            rows.append([
                place,
                f"{user.first_name} {user.last_name or ''}".strip(),
                f"@{user.username}" if user.username else str(user.telegram_id),
                points,
                level,
                _label(user.role, ROLE_LABELS),
                _label(user.participation_status, STATUS_LABELS),
                user.city or "",
            ])
        append_sheet("Активность участников", ["Место", "Участник", "Telegram", "Баллы", "Уровень активности", "Роль", "Статус", "Город"], rows)

    if "users" in sections:
        append_sheet("Участники", ["Имя", "Фамилия", "Telegram", "Дата рождения", "Возраст", "Город", "Email", "Роль", "Статус участия", "Статус заявки", "Баллы", "Дата регистрации"], [[
            u.first_name, u.last_name or "", f"@{u.username}" if u.username else str(u.telegram_id),
            getattr(u, "birth_date", None), u.age, u.city or "", u.email or "",
            _label(u.role, ROLE_LABELS), _label(u.participation_status, STATUS_LABELS),
            _label(u.application_status, APPLICATION_STATUS_LABELS), point_totals.get(u.id, 0), u.created_at,
        ] for u in users])

    if "departments" in sections:
        append_sheet("Департаменты", ["Департамент", "Участников", "Доля от сообщества, %", "Активных целей", "Выполнено целей", "Выполнение целей, %"], [[
            _value(item, "name"), _value(item, "members", 0), _safe_percent(_value(item, "members", 0), len(users)),
            _value(item, "active_goals", 0), _value(item, "done_goals", 0),
            _safe_percent(_value(item, "done_goals", 0), _value(item, "active_goals", 0) + _value(item, "done_goals", 0)),
        ] for item in department_stats])

    if "directions" in sections:
        append_sheet("Направления", ["Департамент", "Направление", "Участников", "Доля от сообщества, %"], [[
            _value(item, "department"), _value(item, "name"), _value(item, "members", 0), _safe_percent(_value(item, "members", 0), len(users)),
        ] for item in direction_stats])

    if "events" in sections:
        append_sheet("Мероприятия", ["Название", "Дата", "Время", "Место", "Формат", "Статус", "Лимит", "Баллы", "Проект ID", "Ответственный ID"], [[
            e.title, e.event_date, e.event_time, e.location, e.format, _label(e.status, EVENT_STATUS_LABELS),
            e.participant_limit, e.points_for_visit, getattr(e, "project_id", None), getattr(e, "responsible_id", None),
        ] for e in events])
        append_sheet("Статусы мероприятий", ["Статус", "Количество", "Доля, %"], [[name, count, _safe_percent(count, len(events))] for name, count in event_status_counts.items()])

    if "projects" in sections:
        append_sheet("Проекты", ["Название", "Автор ID", "Статус", "Площадка", "Дата", "Время", "Создан"], [[
            p.title, p.author_id, _label(p.status, PROJECT_STATUS_LABELS), p.venue_status, p.proposed_date, p.proposed_time, p.created_at,
        ] for p in projects])
        append_sheet("Статусы проектов", ["Статус", "Количество", "Доля, %"], [[name, count, _safe_percent(count, len(projects))] for name, count in project_status_counts.items()])

    if "goals" in sections:
        append_sheet("Цели месяца", ["Месяц", "Уровень", "Название", "План", "Факт", "Выполнение, %", "Статус", "Срок"], [[
            _value(g, "month"), _value(g, "scope_name", _value(g, "scope_type")), _value(g, "title"),
            _value(g, "target_value", 0), _value(g, "current_value", 0),
            _safe_percent(_value(g, "current_value", 0), _value(g, "target_value", 0)), _value(g, "status"), _value(g, "due_date"),
        ] for g in goals])

    if "contacts" in sections:
        append_sheet("Организации", ["Организация", "Контакт", "Должность", "Второй контакт", "Должность 2", "Почта", "Телефон", "Заметки"], [[
            c.organization_name, c.contact_name or "", c.position or "", c.second_contact_name or "", c.second_position or "", c.email or "", c.phone or "", c.notes or "",
        ] for c in contacts])

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
