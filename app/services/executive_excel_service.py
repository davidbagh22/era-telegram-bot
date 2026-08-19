from __future__ import annotations

from collections import Counter
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database.models import (
    EventRegistration,
    LeadershipAttentionItem,
    PortfolioItem,
    ProjectMember,
    RewardRedemption,
    Task,
    TaskSubmission,
    User,
)
from app.database.participation_models import ParticipationLifecycle
from app.utils.constants import ApplicationStatus


SHEET_NAMES = [
    "01 · Сводка",
    "02 · Люди",
    "03 · Рост",
    "04 · Департаменты",
    "05 · Направления",
    "06 · Мой вектор",
    "07 · События",
    "08 · Проекты",
    "09 · Задания",
    "10 · Возможности и документы",
    "11 · Портфолио и признание",
    "12 · Решения",
    "13 · Методика",
]

ERA_DARK = "111216"
ERA_RED = "E32636"
ERA_CREAM = "F3E8D5"
ERA_LIGHT = "F7F4EF"


def resolve_report_period(
    period: str,
    *,
    start_date: date | None = None,
    end_date: date | None = None,
    today: date | None = None,
) -> tuple[date, date, str]:
    end = end_date or today or datetime.now(timezone.utc).date()
    if period == "custom":
        if start_date is None or end_date is None or start_date > end_date:
            raise ValueError("invalid_custom_period")
        return start_date, end_date, f"{start_date:%d.%m.%Y}–{end_date:%d.%m.%Y}"
    days = {"30d": 30, "3m": 90, "6m": 180, "1y": 365}.get(period)
    if days is None:
        raise ValueError("invalid_report_period")
    start = end - timedelta(days=days - 1)
    return start, end, {"30d": "30 дней", "3m": "3 месяца", "6m": "6 месяцев", "1y": "1 год"}[period]


def _date_of(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _in_period(value: Any, start: date, end: date) -> bool:
    day = _date_of(value)
    return day is not None and start <= day <= end


def _enum(value: Any) -> str:
    raw = getattr(value, "value", value)
    return str(raw or "—")


def _name(user: User | None) -> str:
    if user is None:
        return "—"
    return " ".join(part for part in (user.first_name, user.last_name) if part).strip() or "Участник"


def _telegram(user: User | None) -> str:
    if user is None:
        return "—"
    return f"@{user.username}" if user.username else "—"


def _append_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill("solid", fgColor=ERA_DARK)
        cell.font = Font(name="Aptos", bold=True, color=ERA_CREAM)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    if not rows:
        rows = [["Данных за выбранный период нет"] + [""] * (len(headers) - 1)]
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    for col in range(1, ws.max_column + 1):
        values = [str(ws.cell(row, col).value or "") for row in range(1, min(ws.max_row, 100) + 1)]
        ws.column_dimensions[get_column_letter(col)].width = min(max(max(map(len, values), default=0) + 2, 11), 42)


def _summary_sheet(ws, *, period_label: str, rows: list[list[Any]]) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"] = "ERA PLATFORM · EXECUTIVE REPORT"
    ws["A1"].fill = PatternFill("solid", fgColor=ERA_DARK)
    ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=ERA_CREAM)
    ws["A2"] = f"Период: {period_label}"
    ws["A2"].font = Font(name="Aptos", italic=True, color="666666")
    ws.append([])
    ws.append(["Показатель", "Значение", "Что означает", "Источник"])
    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor=ERA_RED)
        cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 32


async def build_executive_workbook(
    session: AsyncSession,
    *,
    analytics,
    health,
    efficiency,
    start_date: date,
    end_date: date,
    period_label: str,
) -> bytes:
    """Build the single executive workbook defined by MASTER section 54.

    It deliberately contains no database/Telegram technical IDs and no raw My
    Vector answers. Vector data enters only through the already cohort-suppressed
    organization-health aggregate supplied by ``build_extended_organization_health``.
    """

    users = list(analytics.users)
    user_by_id = {user.id: user for user in users}
    approved = [u for u in users if u.application_status == ApplicationStatus.APPROVED]

    lifecycles = list((await session.scalars(select(ParticipationLifecycle))).all())
    lifecycle_by_user = {item.user_id: item for item in lifecycles}
    current_roster = [
        u for u in approved
        if not u.is_archived
        and not u.is_blocked
        and getattr(lifecycle_by_user.get(u.id), "participation_mode", "ACTIVE") != "EXITED"
    ]
    active_base = [
        u for u in current_roster
        if getattr(lifecycle_by_user.get(u.id), "activity_state", None) == "ACTIVE"
        and getattr(lifecycle_by_user.get(u.id), "participation_mode", "ACTIVE") in {"ACTIVE", "LIGHT"}
    ]
    new_users = [u for u in approved if _in_period(u.created_at, start_date, end_date)]
    returned = [
        u for u in approved
        if _in_period(getattr(lifecycle_by_user.get(u.id), "returned_at", None), start_date, end_date)
    ]

    registrations = list((await session.scalars(select(EventRegistration))).all())
    task_submissions = list((await session.scalars(select(TaskSubmission))).all())
    project_members = list((await session.scalars(select(ProjectMember))).all())
    tasks = list((await session.scalars(select(Task))).all())
    portfolio = list((await session.scalars(select(PortfolioItem))).all())
    rewards = list((await session.scalars(select(RewardRedemption))).all())
    decisions = list((await session.scalars(select(LeadershipAttentionItem))).all())

    period_events = [event for event in analytics.events if _in_period(event.event_date, start_date, end_date)]
    event_ids = {event.id for event in period_events}
    period_regs = [row for row in registrations if row.event_id in event_ids]
    attended = [row for row in period_regs if _enum(row.status) == "attended"]
    period_projects = [project for project in analytics.projects if _in_period(project.created_at, start_date, end_date)]
    period_project_ids = {project.id for project in period_projects}
    period_tasks = [task for task in tasks if _in_period(task.created_at, start_date, end_date)]
    completed_task_submissions = [
        row for row in task_submissions
        if _in_period(row.created_at, start_date, end_date) and _enum(row.status) in {"approved", "completed"}
    ]

    wb = Workbook()
    wb.remove(wb.active)
    for title in SHEET_NAMES:
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = False

    summary_rows = [
        ["Подтверждено исторически", len(approved), "Все когда-либо одобренные профили в доступной базе", "User.application_status"],
        ["Текущий состав", len(current_roster), "Одобрены, не archived/blocked и не EXITED", "ParticipationLifecycle"],
        ["Активная база", len(active_base), "Meaningful Activity + ACTIVE/LIGHT mode; digital points сюда не входят", "ParticipationLifecycle"],
        ["Новые за период", len(new_users), "Одобренные участники, зарегистрированные в периоде", "User.created_at"],
        ["Вернулись за период", len(returned), "Зафиксированный возврат после снижения активности", "ParticipationLifecycle.returned_at"],
        ["События за период", len(period_events), "События с датой внутри выбранного периода", "Event.event_date"],
        ["Attendance rate", round(len(attended) * 100 / len(period_regs), 1) if period_regs else 0, "Подтверждённые посещения / регистрации выбранных событий", "EventRegistration"],
        ["Проекты за период", len(period_projects), "Проекты, созданные в выбранном периоде", "Project.created_at"],
        ["Verified task results", len(completed_task_submissions), "Только подтверждённые результаты задач", "TaskSubmission"],
        ["Эффективность", efficiency.score, efficiency.label, "ERA efficiency service"],
        ["Пульс Мой вектор", health.pulse if health.pulse is not None else "—", health.pulse_label, "Безопасный агрегат; cohort suppression"],
    ]
    _summary_sheet(wb[SHEET_NAMES[0]], period_label=period_label, rows=summary_rows)

    people_rows = []
    for user in current_roster:
        lifecycle = lifecycle_by_user.get(user.id)
        people_rows.append([
            _name(user),
            _telegram(user),
            user.city or "—",
            _enum(user.participation_status),
            getattr(lifecycle, "participation_mode", "ACTIVE"),
            getattr(lifecycle, "activity_state", "ADAPTATION"),
            _date_of(user.created_at),
            _date_of(getattr(lifecycle, "last_meaningful_at", None)),
        ])
    _append_table(wb[SHEET_NAMES[1]], ["Участник", "Telegram", "Город", "Ранг", "Режим", "Состояние", "В ЭРА с", "Последнее meaningful действие"], people_rows)

    rank_counts = Counter(_enum(user.participation_status) for user in current_roster)
    state_counts = Counter(getattr(lifecycle_by_user.get(user.id), "activity_state", "ADAPTATION") for user in current_roster)
    growth_rows = [["Ранг", key, value] for key, value in rank_counts.items()]
    growth_rows += [["Activity State", key, value] for key, value in state_counts.items()]
    growth_rows += [["Период", "Новые", len(new_users)], ["Период", "Вернулись", len(returned)]]
    _append_table(wb[SHEET_NAMES[2]], ["Срез", "Этап", "Количество"], growth_rows)

    _append_table(
        wb[SHEET_NAMES[3]],
        ["Департамент", "Участников", "Активных целей", "Выполнено целей"],
        [[row.name, row.members, row.active_goals, row.done_goals] for row in analytics.department_stats],
    )
    _append_table(
        wb[SHEET_NAMES[4]],
        ["Департамент", "Направление", "Участников"],
        [[row.department, row.name, row.members] for row in analytics.direction_stats],
    )

    vector_rows = [
        ["Coverage", health.pulse_coverage, f"n={health.pulse_sample_size}", "minimum cohort применяется до расчёта"],
        ["Pulse", health.pulse if health.pulse is not None else "—", health.pulse_label, "не используется для rank/certificates/leadership decisions"],
    ]
    for item in health.vector_dimensions:
        vector_rows.append([item.label, item.value, item.delta if item.delta is not None else "—", "агрегат; не индивидуальная оценка"])
    _append_table(wb[SHEET_NAMES[5]], ["Показатель", "Значение", "Изменение / выборка", "Privacy note"], vector_rows)

    reg_counts: Counter[int] = Counter(row.event_id for row in period_regs)
    attended_counts: Counter[int] = Counter(row.event_id for row in attended)
    event_rows = []
    for event in period_events:
        total = reg_counts.get(event.id, 0)
        came = attended_counts.get(event.id, 0)
        event_rows.append([event.title, event.event_date, _enum(event.status), total, came, round(came * 100 / total, 1) if total else 0])
    _append_table(wb[SHEET_NAMES[6]], ["Событие", "Дата", "Статус", "Регистраций", "Посетили", "Attendance, %"], event_rows)

    contributor_counts: Counter[int] = Counter(
        member.project_id
        for member in project_members
        if member.project_id in period_project_ids and getattr(member, "contribution_status", None) == "confirmed"
    )
    _append_table(
        wb[SHEET_NAMES[7]],
        ["Проект", "Статус", "Создан", "Verified contributors"],
        [[project.title, _enum(project.status), _date_of(project.created_at), contributor_counts.get(project.id, 0)] for project in period_projects],
    )

    task_rows = []
    for task in period_tasks:
        task_rows.append([task.title, _enum(task.status), _date_of(task.created_at), _date_of(task.deadline), getattr(task, "points", 0), "командное" if getattr(task, "team_mode", None) == "TEAM" else "обычное"])
    _append_table(wb[SHEET_NAMES[8]], ["Задание", "Статус", "Создано", "Дедлайн", "Баллы", "Тип"], task_rows)

    reward_rows = []
    for row in rewards:
        if not _in_period(row.created_at, start_date, end_date):
            continue
        reward_rows.append([_name(user_by_id.get(row.user_id)), _enum(row.status), _date_of(row.created_at), "Заявка на возможность / документ"])
    _append_table(wb[SHEET_NAMES[9]], ["Участник", "Статус", "Дата", "Тип записи"], reward_rows)

    portfolio_rows = []
    for item in portfolio:
        if not _in_period(item.created_at, start_date, end_date) and not _in_period(getattr(item, "issued_at", None), start_date, end_date):
            continue
        portfolio_rows.append([_name(user_by_id.get(item.user_id)), item.title, _enum(item.item_type), _enum(item.status), _date_of(getattr(item, "issued_at", None) or item.created_at)])
    _append_table(wb[SHEET_NAMES[10]], ["Участник", "Достижение", "Тип", "Статус", "Дата"], portfolio_rows)

    decision_rows = []
    for item in decisions:
        if not _in_period(item.created_at, start_date, end_date):
            continue
        decision_rows.append([_enum(item.type), _enum(item.severity), _enum(item.scope_type), _enum(item.status), getattr(item, "resolution", None) or "—", _date_of(item.created_at)])
    _append_table(wb[SHEET_NAMES[11]], ["Сигнал / решение", "Критичность", "Scope", "Статус", "Решение", "Дата"], decision_rows)

    method_rows = [
        ["Период отчёта", period_label, "Фильтр применяется к временным сущностям; current roster/Active Base показываются на дату формирования."],
        ["Active Base", "Meaningful Activity + ACTIVE/LIGHT", "Не считается по login, просмотру, daily points, регистрации, должности или нахождению в департаменте."],
        ["Ранг", "ParticipationStatus", "Ранг ≠ должность и не определяется только total points."],
        ["Verified result", "Только подтверждённый результат", "Task/Event/Project scoring идёт через единый idempotent verified-activity gateway."],
        ["Мой вектор", "Aggregate only", "Raw ответы, personal notes и индивидуальные психологические оценки в workbook не попадают."],
        ["Technical IDs", "Не экспортируются", "В таблицах используются названия/имена; database ids и Telegram numeric ids скрыты по умолчанию."],
        ["PII", "Минимизация", "Телефон/email не входят в executive workbook; отдельные операционные exports имеют свой permission scope."],
    ]
    _append_table(wb[SHEET_NAMES[12]], ["Правило", "Как считается", "Ограничение / смысл"], method_rows)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
