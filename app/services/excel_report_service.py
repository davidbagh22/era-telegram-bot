from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ERA_RED = "E32636"
ERA_RED_DARK = "981B28"
ERA_DARK = "111216"
ERA_SURFACE = "1B1C21"
ERA_CREAM = "F3E8D5"
ERA_MUTED = "A8A9B0"
ERA_LIGHT = "F7F4EF"
ERA_BORDER = "D9D2C7"
ERA_GREEN = "3A8F63"
ERA_AMBER = "B9842B"


def _bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _is_header_row(values: list[Any]) -> bool:
    non_empty = [value for value in values if value not in (None, "")]
    if len(non_empty) < 2:
        return False
    return all(isinstance(value, str) for value in non_empty)


def polish_workbook(content: bytes) -> bytes:
    """Apply one concise ERA visual language to any generated XLSX.

    This deliberately avoids decorative data dumps: raw technical ID columns
    that have no management value are removed, widths are bounded, gridlines
    are hidden and readable filters/frozen headers are kept where possible.
    """

    wb = load_workbook(BytesIO(content))
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_view.zoomScale = 90

        # Remove internal implementation identifiers from management exports.
        removable = {"Проект ID", "Ответственный ID", "Автор ID", "scope_id", "user_id"}
        header_row = None
        for row_idx in range(1, min(ws.max_row, 8) + 1):
            values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
            if _is_header_row(values) and any(str(value) in removable for value in values if value is not None):
                header_row = row_idx
                for col in range(ws.max_column, 0, -1):
                    if str(ws.cell(row_idx, col).value or "") in removable:
                        ws.delete_cols(col, 1)
                break

        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name="Aptos", size=10.5, color="242126", bold=cell.font.bold)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.row % 2 == 0 and cell.fill.fill_type is None:
                    cell.fill = PatternFill("solid", fgColor="FCFAF7")

        # Style likely table headers, while preserving dedicated title rows.
        for row_idx in range(1, min(ws.max_row, 12) + 1):
            values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
            if not _is_header_row(values):
                continue
            if len([v for v in values if v not in (None, "")]) < 2:
                continue
            for cell in ws[row_idx]:
                if cell.value in (None, ""):
                    continue
                cell.fill = PatternFill("solid", fgColor=ERA_DARK)
                cell.font = Font(name="Aptos", size=10.5, bold=True, color=ERA_CREAM)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = Border(bottom=Side(style="thin", color=ERA_RED))
            if ws.freeze_panes is None:
                ws.freeze_panes = f"A{row_idx + 1}"
            break

        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            values = [str(ws.cell(row, col).value or "") for row in range(1, min(ws.max_row, 80) + 1)]
            longest = max((len(value) for value in values), default=0)
            ws.column_dimensions[letter].width = min(max(longest + 2, 11), 38)

        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}" if ws.cell(1, 1).value else ws.auto_filter.ref

        ws.row_dimensions[1].height = max(ws.row_dimensions[1].height or 15, 24)

    return _bytes(wb)


def _title(ws, title: str, subtitle: str | None = None, *, columns: int = 5) -> int:
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    ws["A1"] = title
    ws["A1"].fill = PatternFill("solid", fgColor=ERA_DARK)
    ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=ERA_CREAM)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
        ws["A2"] = subtitle
        ws["A2"].font = Font(name="Aptos", size=10, color="6E6870")
        ws["A2"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 30
        return 4
    return 3


def _section_header(ws, row: int, title: str, *, columns: int = 5) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
    cell = ws.cell(row, 1, title)
    cell.fill = PatternFill("solid", fgColor=ERA_RED)
    cell.font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22
    return row + 1


def _table_header(ws, row: int, labels: list[str]) -> int:
    for col, label in enumerate(labels, 1):
        cell = ws.cell(row, col, label)
        cell.fill = PatternFill("solid", fgColor=ERA_DARK)
        cell.font = Font(name="Aptos", bold=True, color=ERA_CREAM)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color=ERA_RED))
    return row + 1


def add_health_sheets(content: bytes, health, efficiency) -> bytes:
    """Put an executive organization-health view at the front of a workbook."""

    wb = load_workbook(BytesIO(content))
    for name in ("Здоровье организации", "Пульс Мой вектор", "Все показатели"):
        if name in wb.sheetnames:
            del wb[name]

    ws = wb.create_sheet("Здоровье организации", 0)
    row = _title(
        ws,
        "ЗДОРОВЬЕ ОРГАНИЗАЦИИ · ЭРА",
        "Два независимых сигнала: эффективность работы и безопасный агрегированный Пульс сообщества.",
    )
    ws["A4"] = "Эффективность ЭРА"
    ws["B4"] = efficiency.score
    ws["C4"] = "/100"
    ws["D4"] = efficiency.label
    ws["A5"] = "Пульс организации"
    ws["B5"] = health.pulse if health.pulse is not None else "—"
    ws["C5"] = "/100" if health.pulse is not None else ""
    ws["D5"] = health.pulse_label
    ws["A6"] = "Охват Пульса"
    ws["B6"] = health.pulse_coverage / 100
    ws["C6"] = f"{health.pulse_sample_size} Check-in"
    ws["D6"] = "показывается только при безопасной выборке ≥5"
    ws["B6"].number_format = "0%"
    for r in (4, 5, 6):
        ws.cell(r, 1).font = Font(name="Aptos", bold=True, color=ERA_DARK)
        ws.cell(r, 2).font = Font(name="Aptos Display", size=18 if r < 6 else 12, bold=True, color=ERA_RED_DARK)
        ws.cell(r, 4).font = Font(name="Aptos", color="5E5860")

    row = 8
    row = _section_header(ws, row, "Сигналы внимания")
    if health.risks:
        for risk in health.risks:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.cell(row, 1, f"• {risk}")
            ws.cell(row, 1).fill = PatternFill("solid", fgColor="FFF4E5")
            ws.cell(row, 1).font = Font(name="Aptos", color="6C4A15")
            ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row, 1, "Критических сигналов по доступным данным нет.")
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="EAF6EF")
        ws.cell(row, 1).font = Font(name="Aptos", color="245D43")
        row += 1

    row += 1
    row = _section_header(ws, row, "Ключевые показатели")
    row = _table_header(ws, row, ["Блок", "Показатель", "Значение", "Что показывает", "Оценка / 100"])
    executive_keys = {
        "active_30d", "retention_30d", "growth_conversion", "attendance_rate", "feedback",
        "active_projects", "completed_tasks_30d", "overdue_tasks", "queue", "task_delivery",
    }
    for metric in health.metrics:
        if metric.key not in executive_keys:
            continue
        ws.append([metric.category, metric.label, metric.display, metric.note, metric.score if metric.score is not None else "—"])
        row += 1
    ws.freeze_panes = "A13"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 31
    ws.column_dimensions["C"].width = 17
    ws.column_dimensions["D"].width = 52
    ws.column_dimensions["E"].width = 15

    pulse_ws = wb.create_sheet("Пульс Мой вектор", 1)
    row = _title(
        pulse_ws,
        "ПУЛЬС ОРГАНИЗАЦИИ",
        health.data_note,
        columns=4,
    )
    pulse_ws["A4"] = "Пульс"
    pulse_ws["B4"] = health.pulse if health.pulse is not None else "Недостаточно данных"
    pulse_ws["A5"] = "Охват"
    pulse_ws["B5"] = health.pulse_coverage / 100
    pulse_ws["B5"].number_format = "0%"
    pulse_ws["C5"] = f"n={health.pulse_sample_size}"
    row = 7
    row = _table_header(pulse_ws, row, ["Область", "Текущее состояние", "Изменение", "Комментарий"])
    if health.vector_dimensions:
        for item in health.vector_dimensions:
            pulse_ws.append([
                item.label,
                item.value,
                item.delta if item.delta is not None else "—",
                "агрегат текущего состояния; не характеристика личности",
            ])
    else:
        pulse_ws.append(["—", "—", "—", "Недостаточно Check-in для безопасной групповой аналитики"])
    pulse_ws.freeze_panes = "A8"
    pulse_ws.column_dimensions["A"].width = 24
    pulse_ws.column_dimensions["B"].width = 22
    pulse_ws.column_dimensions["C"].width = 16
    pulse_ws.column_dimensions["D"].width = 55

    if health.vector_dimensions:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Пять областей Пульса"
        chart.x_axis.title = "0–100"
        chart.height = 7
        chart.width = 12
        chart.add_data(Reference(pulse_ws, min_col=2, min_row=7, max_row=7 + len(health.vector_dimensions)), titles_from_data=True)
        chart.set_categories(Reference(pulse_ws, min_col=1, min_row=8, max_row=7 + len(health.vector_dimensions)))
        pulse_ws.add_chart(chart, "F7")

    metrics_ws = wb.create_sheet("Все показатели", 2)
    row = _title(
        metrics_ws,
        "ПОКАЗАТЕЛИ ЗДОРОВЬЯ ЭРА",
        health.period_label,
    )
    row = _table_header(metrics_ws, row, ["Блок", "Показатель", "Значение", "Что означает", "Оценка / 100"])
    for metric in health.metrics:
        metrics_ws.append([
            metric.category,
            metric.label,
            metric.display,
            metric.note,
            metric.score if metric.score is not None else "—",
        ])
    metrics_ws.freeze_panes = "A5"
    metrics_ws.auto_filter.ref = f"A4:E{metrics_ws.max_row}"
    metrics_ws.column_dimensions["A"].width = 18
    metrics_ws.column_dimensions["B"].width = 34
    metrics_ws.column_dimensions["C"].width = 18
    metrics_ws.column_dimensions["D"].width = 58
    metrics_ws.column_dimensions["E"].width = 15

    return polish_workbook(_bytes(wb))


def build_development_workbook(result: dict[str, Any]) -> bytes:
    """Compact privacy-safe XLSX for the My Vector aggregate export."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Пульс Мой вектор"
    _title(
        ws,
        "МОЙ ВЕКТОР · СВОДКА ДЛЯ ЭРА",
        "Только безопасные агрегаты. Личные заметки, сырые ответы и индивидуальные профили не выгружаются.",
        columns=4,
    )
    ws.append(["Охват", result.get("coverage_percent", 0) / 100, f"n={result.get('sample_size', 0)}", f"минимальная выборка: {result.get('minimum_cohort', 5)}"])
    ws["B4"].number_format = "0%"
    if result.get("suppressed"):
        ws.append(["Пульс", "—", "", "Недостаточно ответов для безопасной групповой аналитики"])
    else:
        ws.append(["Пульс", result.get("index", 0), "/100", "агрегат пяти текущих состояний"])
        row = 7
        row = _table_header(ws, row, ["Область", "Значение", "Изменение", "Охват"])
        from app.services import development_service as dev

        state = result.get("state") or {}
        delta = result.get("delta") or {}
        for code in dev.STATE_DIMENSIONS:
            ws.append([dev.STATE_LABELS[code], state.get(code, 0), delta.get(code, "—"), result.get("coverage_percent", 0) / 100])
            ws.cell(ws.max_row, 4).number_format = "0%"

        wants = result.get("development_wants") or []
        if wants:
            ws2 = wb.create_sheet("Запросы на развитие")
            _title(ws2, "ЗАПРОСЫ НА РАЗВИТИЕ", "Сводные добровольные ответы", columns=3)
            _table_header(ws2, 4, ["Запрос", "Доля", "Ответов"])
            for item in wants:
                ws2.append([item.get("key"), (item.get("percent") or 0) / 100, item.get("count")])
                ws2.cell(ws2.max_row, 2).number_format = "0%"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 52
    return polish_workbook(_bytes(wb))
