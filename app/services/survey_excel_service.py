from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database.management_models import AdminSurvey, AdminSurveyResponse
from app.database.models import User
from app.services.survey_service import answer_items, survey_questions

HEADER_FILL = PatternFill("solid", fgColor="7B2FF7")
LIGHT_FILL = PatternFill("solid", fgColor="F5F1FF")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="241332")


def _value(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y %H:%M")
    return str(value)


def _setup_sheet(ws, title: str, headers: list[str]) -> None:
    ws.title = title[:31]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _finish_sheet(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in ws.columns:
        max_len = max(len(_value(cell.value)) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_len + 3, 14), 55)


def build_survey_workbook(
    surveys: Iterable[AdminSurvey],
    responses: Iterable[AdminSurveyResponse],
    users: Iterable[User] = (),
) -> bytes:
    surveys = list(surveys)
    responses = list(responses)
    user_map = {user.id: user for user in users}
    survey_map = {survey.id: survey for survey in surveys}

    wb = Workbook()
    summary = wb.active
    summary.title = "Сводка"
    summary.append(["Показатель", "Значение"])
    summary.append(["Опросов", len(surveys)])
    summary.append(["Ответов", len(responses)])
    summary.append(["Активных/отправленных", sum(1 for s in surveys if s.status in {"draft", "sent", "active"})])
    summary.append(["Ежемесячных", sum(1 for s in surveys if s.is_monthly)])
    summary["A1"].font = TITLE_FONT
    for cell in summary[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
    _finish_sheet(summary)

    ws = wb.create_sheet("Опросы")
    _setup_sheet(ws, "Опросы", ["ID", "Название", "Статус", "Ежемесячный", "Вопросов", "Ответов", "Последняя рассылка", "Месяц", "Описание"])
    response_counts: dict[int, int] = {}
    for response in responses:
        response_counts[response.survey_id] = response_counts.get(response.survey_id, 0) + 1
    for survey in surveys:
        ws.append([
            survey.id,
            survey.title,
            survey.status,
            "Да" if survey.is_monthly else "Нет",
            len(survey_questions(survey)),
            response_counts.get(survey.id, 0),
            _value(survey.sent_at),
            survey.last_sent_month or "",
            survey.description or "",
        ])
    _finish_sheet(ws)

    ws = wb.create_sheet("Ответы")
    _setup_sheet(ws, "Ответы", ["Опрос", "Участник", "Telegram", "Дата ответа", "Вопрос", "Ответ"])
    for response in responses:
        survey = survey_map.get(response.survey_id)
        user = user_map.get(response.user_id)
        user_name = ""
        telegram = ""
        if user:
            user_name = f"{user.first_name} {user.last_name or ''}".strip()
            telegram = f"@{user.username}" if user.username else str(user.telegram_id)
        for item in answer_items(response):
            ws.append([
                survey.title if survey else response.survey_id,
                user_name or response.user_id,
                telegram,
                _value(response.submitted_at or response.created_at),
                item["question"],
                item["answer"],
            ])
    _finish_sheet(ws)

    ws = wb.create_sheet("По вопросам")
    _setup_sheet(ws, "По вопросам", ["Опрос", "Вопрос", "Количество ответов", "Последние ответы"])
    grouped: dict[tuple[int, str], list[str]] = {}
    for response in responses:
        for item in answer_items(response):
            key = (response.survey_id, item["question"])
            grouped.setdefault(key, []).append(item["answer"])
    for (survey_id, question), answers in grouped.items():
        survey = survey_map.get(survey_id)
        ws.append([
            survey.title if survey else survey_id,
            question,
            len(answers),
            "\n".join(answers[-7:]),
        ])
    _finish_sheet(ws)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows(min_row=2):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = LIGHT_FILL

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
