from __future__ import annotations

import csv
from enum import Enum
from io import BytesIO, StringIO
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Response
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user, get_session, get_settings
from app.config import Settings
from app.database.models import User
from app.services.admin_analytics_service import build_analytics_payload
from app.services.admin_dashboard_service import has_dashboard_access
from app.services.era_efficiency_service import build_efficiency_snapshot
from app.services.excel_quality_service import finalize_business_workbook
from app.services.excel_report_service import add_health_sheets, polish_workbook
from app.services.excel_service import build_analytics_workbook
from app.services.organization_health_extended_service import build_extended_organization_health

router = APIRouter(prefix="/admin/analytics", tags=["admin-analytics-details"])

AnalyticsDetailSection = Literal["users", "events", "projects", "contacts", "goals"]


async def require_dashboard_access(
    user: User = Depends(get_current_user),
    settings: Settings = Depends(get_settings),
) -> User:
    if not has_dashboard_access(user, settings, user.telegram_id):
        raise HTTPException(status_code=403, detail="admin_access_required")
    return user


class AnalyticsDetailItemOut(BaseModel):
    id: int
    title: str
    subtitle: str | None = None
    status: str | None = None


class AnalyticsDetailsOut(BaseModel):
    section: AnalyticsDetailSection
    total: int
    items: list[AnalyticsDetailItemOut]


class EfficiencyMetricOut(BaseModel):
    key: str
    label: str
    value: float
    display: str
    score: int | None
    note: str


class EfficiencyRecommendationOut(BaseModel):
    priority: str
    title: str
    reason: str
    action: str


class EfficiencyOut(BaseModel):
    score: int
    label: str
    period_label: str
    metrics: list[EfficiencyMetricOut]
    recommendations: list[EfficiencyRecommendationOut]
    top_interest: str | None
    top_interest_count: int
    data_note: str


class HealthMetricOut(BaseModel):
    key: str
    category: str
    label: str
    value: float
    display: str
    note: str
    score: int | None = None


class VectorDimensionOut(BaseModel):
    key: str
    label: str
    value: int
    delta: int | None = None


class OrganizationHealthOut(BaseModel):
    pulse: int | None
    pulse_label: str
    pulse_coverage: int
    pulse_sample_size: int
    pulse_suppressed: bool
    vector_dimensions: list[VectorDimensionOut]
    metrics: list[HealthMetricOut]
    risks: list[str]
    period_label: str
    data_note: str


def _text(value: object | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, Enum):
        value = value.value
    rendered = str(value).strip()
    return rendered or None


def _detail_items(section: AnalyticsDetailSection, data) -> list[AnalyticsDetailItemOut]:
    if section == "users":
        return [
            AnalyticsDetailItemOut(
                id=user.id,
                title=" ".join(part for part in [user.first_name, user.last_name] if part).strip()
                or f"Участник #{user.id}",
                subtitle=" · ".join(
                    part
                    for part in [
                        _text(user.role),
                        f"@{user.username}" if user.username else None,
                    ]
                    if part
                )
                or None,
                status=_text(user.application_status),
            )
            for user in data.users
        ]
    if section == "events":
        return [
            AnalyticsDetailItemOut(
                id=event.id,
                title=event.title,
                subtitle=" · ".join(
                    part
                    for part in [
                        event.event_date.isoformat() if event.event_date else None,
                        event.event_time.strftime("%H:%M") if event.event_time else None,
                    ]
                    if part
                )
                or None,
                status=_text(event.status),
            )
            for event in data.events
        ]
    if section == "projects":
        return [
            AnalyticsDetailItemOut(
                id=project.id,
                title=project.title,
                subtitle=project.created_at.date().isoformat() if project.created_at else None,
                status=_text(project.status),
            )
            for project in data.projects
        ]
    if section == "contacts":
        return [
            AnalyticsDetailItemOut(
                id=contact.id,
                title=contact.organization_name,
                subtitle=None,
                status="active",
            )
            for contact in data.contacts
        ]
    return [
        AnalyticsDetailItemOut(
            id=goal.id,
            title=goal.title,
            subtitle=f"{goal.scope_name} · {goal.current_value}/{goal.target_value}",
            status=goal.status,
        )
        for goal in data.goals
    ]


def _with_efficiency_sheet(content: bytes, snapshot) -> bytes:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = load_workbook(BytesIO(content))
    if "Эффективность" in wb.sheetnames:
        del wb["Эффективность"]
    ws = wb.create_sheet("Эффективность", 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "ЭФФЕКТИВНОСТЬ ЭРА"
    ws["A1"].font = Font(size=18, bold=True, color="F3E8D5")
    ws["A1"].fill = PatternFill("solid", fgColor="111216")
    ws.merge_cells("A1:D1")
    ws["A2"] = f"{snapshot.score}/100 · {snapshot.label}"
    ws["A2"].font = Font(size=14, bold=True, color="981B28")
    ws.merge_cells("A2:D2")
    ws["A3"] = snapshot.data_note
    ws.merge_cells("A3:D3")

    ws.append([])
    ws.append(["Показатель", "Значение", "Оценка / 100", "Что это значит"])
    for metric in snapshot.metrics:
        ws.append([metric.label, metric.display, metric.score if metric.score is not None else "—", metric.note])

    ws.append([])
    ws.append(["ПРИОРИТЕТЫ НА ЭТУ НЕДЕЛЮ", "Почему", "Что сделать", "Приоритет"])
    for item in snapshot.recommendations:
        ws.append([item.title, item.reason, item.action, item.priority])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 24
    ws.freeze_panes = "A5"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


@router.get("/weekly", response_model=EfficiencyOut)
async def read_weekly_efficiency(
    _admin: User = Depends(require_dashboard_access),
    session: AsyncSession = Depends(get_session),
) -> EfficiencyOut:
    snapshot = await build_efficiency_snapshot(session)
    return EfficiencyOut(
        score=snapshot.score,
        label=snapshot.label,
        period_label=snapshot.period_label,
        metrics=[EfficiencyMetricOut(**item.__dict__) for item in snapshot.metrics],
        recommendations=[EfficiencyRecommendationOut(**item.__dict__) for item in snapshot.recommendations],
        top_interest=snapshot.top_interest,
        top_interest_count=snapshot.top_interest_count,
        data_note=snapshot.data_note,
    )


@router.get("/health", response_model=OrganizationHealthOut)
async def read_organization_health(
    _admin: User = Depends(require_dashboard_access),
    session: AsyncSession = Depends(get_session),
) -> OrganizationHealthOut:
    health = await build_extended_organization_health(session)
    return OrganizationHealthOut(
        pulse=health.pulse,
        pulse_label=health.pulse_label,
        pulse_coverage=health.pulse_coverage,
        pulse_sample_size=health.pulse_sample_size,
        pulse_suppressed=health.pulse_suppressed,
        vector_dimensions=[VectorDimensionOut(**item.__dict__) for item in health.vector_dimensions],
        metrics=[HealthMetricOut(**item.__dict__) for item in health.metrics],
        risks=health.risks,
        period_label=health.period_label,
        data_note=health.data_note,
    )


@router.get("/health-report.xlsx")
async def export_organization_health_report(
    _admin: User = Depends(require_dashboard_access),
    session: AsyncSession = Depends(get_session),
) -> Response:
    data = await build_analytics_payload(session)
    efficiency = await build_efficiency_snapshot(session)
    health = await build_extended_organization_health(session)
    base = build_analytics_workbook(
        data.users,
        data.events,
        data.projects,
        data.totals,
        department_stats=data.department_stats,
        direction_stats=data.direction_stats,
        goals=data.goals,
        contacts=data.contacts,
        sections={"summary"},
    )
    content = finalize_business_workbook(add_health_sheets(base, health, efficiency))
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ERA_organization_health.xlsx"'},
    )


@router.get("/full-report.xlsx")
async def export_full_analytics_report(
    _admin: User = Depends(require_dashboard_access),
    session: AsyncSession = Depends(get_session),
) -> Response:
    data = await build_analytics_payload(session)
    efficiency = await build_efficiency_snapshot(session)
    health = await build_extended_organization_health(session)
    base = build_analytics_workbook(
        data.users,
        data.events,
        data.projects,
        data.totals,
        department_stats=data.department_stats,
        direction_stats=data.direction_stats,
        goals=data.goals,
        contacts=data.contacts,
    )
    content = _with_efficiency_sheet(base, efficiency)
    content = finalize_business_workbook(add_health_sheets(content, health, efficiency))
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ERA_full_report.xlsx"'},
    )


@router.get("/details/{section}", response_model=AnalyticsDetailsOut)
async def read_analytics_details(
    section: AnalyticsDetailSection,
    _admin: User = Depends(require_dashboard_access),
    session: AsyncSession = Depends(get_session),
) -> AnalyticsDetailsOut:
    data = await build_analytics_payload(session)
    items = _detail_items(section, data)
    return AnalyticsDetailsOut(section=section, total=len(items), items=items)


@router.get("/details/{section}/export.xlsx")
async def export_analytics_details_xlsx(
    section: AnalyticsDetailSection,
    _admin: User = Depends(require_dashboard_access),
    session: AsyncSession = Depends(get_session),
) -> Response:
    data = await build_analytics_payload(session)
    sections = {
        "users": {"summary", "users"},
        "events": {"summary", "events"},
        "projects": {"summary", "projects"},
        "contacts": {"summary", "contacts"},
        "goals": {"summary", "goals"},
    }[section]
    content = finalize_business_workbook(
        polish_workbook(
            build_analytics_workbook(
                data.users,
                data.events,
                data.projects,
                data.totals,
                department_stats=data.department_stats,
                direction_stats=data.direction_stats,
                goals=data.goals,
                contacts=data.contacts,
                sections=sections,
            )
        )
    )
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="ERA_{section}.xlsx"'},
    )


# Compatibility for older clients; new Mini App surfaces XLSX only.
@router.get("/details/{section}/export.csv")
async def export_analytics_details_csv(
    section: AnalyticsDetailSection,
    _admin: User = Depends(require_dashboard_access),
    session: AsyncSession = Depends(get_session),
) -> Response:
    data = await build_analytics_payload(session)
    items = _detail_items(section, data)
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Название", "Детали", "Статус"])
    for item in items:
        writer.writerow([item.title, item.subtitle or "", item.status or ""])
    payload = "\ufeff" + buffer.getvalue()
    return Response(
        content=payload.encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="ERA_{section}.csv"'},
    )
