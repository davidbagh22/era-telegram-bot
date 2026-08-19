from __future__ import annotations

from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user, get_session, get_settings
from app.config import Settings
from app.database.models import User
from app.services.admin_dashboard_service import has_dashboard_access
from app.services.admin_metric_drilldown_service import METRIC_LABELS, build_metric_drilldown
from app.services.excel_quality_service import finalize_business_workbook

router = APIRouter(prefix="/admin/analytics/drilldown", tags=["admin-analytics-drilldown"])


async def require_full_admin(
    user: User = Depends(get_current_user),
    settings: Settings = Depends(get_settings),
) -> User:
    if not has_dashboard_access(user, settings, user.telegram_id):
        raise HTTPException(status_code=403, detail="admin_access_required")
    return user


class MetricRowOut(BaseModel):
    id: int
    entity_type: str
    entity_id: int
    title: str
    subtitle: str | None = None
    status: str | None = None


class MetricDrilldownOut(BaseModel):
    metric: str
    label: str
    total: int
    items: list[MetricRowOut]


async def _load(session: AsyncSession, metric: str):
    if metric not in METRIC_LABELS:
        raise HTTPException(status_code=404, detail="unknown_admin_metric")
    try:
        return await build_metric_drilldown(session, metric)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@router.get("/{metric}", response_model=MetricDrilldownOut)
async def read_metric_drilldown(
    metric: str,
    _admin: User = Depends(require_full_admin),
    session: AsyncSession = Depends(get_session),
) -> MetricDrilldownOut:
    result = await _load(session, metric)
    return MetricDrilldownOut(
        metric=result.metric,
        label=result.label,
        total=result.total,
        items=[MetricRowOut(**row.__dict__) for row in result.rows],
    )


def _workbook_bytes(label: str, rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"
    ws["A1"] = f"ЭРА · {label}"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:E1")
    ws.append([])
    ws.append(["№", "Название", "Контекст", "Статус", "Тип записи"])
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    for index, row in enumerate(rows, start=1):
        ws.append([index, row.title, row.subtitle or "", row.status or "", row.entity_type])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:E{max(3, ws.max_row)}"
    widths = [8, 36, 54, 22, 24]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    raw = BytesIO()
    wb.save(raw)
    return finalize_business_workbook(raw.getvalue())


@router.get("/{metric}.xlsx")
async def export_metric_drilldown(
    metric: str,
    _admin: User = Depends(require_full_admin),
    session: AsyncSession = Depends(get_session),
) -> Response:
    result = await _load(session, metric)
    content = _workbook_bytes(result.label, result.rows)
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="era-{metric}.xlsx"',
            "X-ERA-Metric": result.metric,
            "X-ERA-Total": str(result.total),
        },
    )
