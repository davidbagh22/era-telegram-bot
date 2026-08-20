from __future__ import annotations

import unittest
from datetime import date
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.database.base import Base
from app.services.executive_excel_service import SHEET_NAMES, build_executive_workbook, resolve_report_period


class ExecutivePeriodTests(unittest.TestCase):
    def test_standard_periods_are_supported(self) -> None:
        today = date(2026, 8, 19)
        self.assertEqual(resolve_report_period("30d", today=today)[0], date(2026, 7, 21))
        self.assertEqual(resolve_report_period("3m", today=today)[2], "3 месяца")
        self.assertEqual(resolve_report_period("6m", today=today)[2], "6 месяцев")
        self.assertEqual(resolve_report_period("1y", today=today)[2], "1 год")

    def test_custom_period_requires_both_dates(self) -> None:
        with self.assertRaises(ValueError):
            resolve_report_period("custom", start_date=date(2026, 1, 1))
        start, end, label = resolve_report_period(
            "custom",
            start_date=date(2026, 1, 1),
            end_date=date(2026, 2, 1),
        )
        self.assertEqual((start, end), (date(2026, 1, 1), date(2026, 2, 1)))
        self.assertIn("01.01.2026", label)


class ExecutiveWorkbookTests(unittest.IsolatedAsyncioTestCase):
    async def asyncSetUp(self) -> None:
        self.engine = create_async_engine("sqlite+aiosqlite:///:memory:")
        async with self.engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)
        self.session_factory = async_sessionmaker(self.engine, expire_on_commit=False)

    async def asyncTearDown(self) -> None:
        await self.engine.dispose()

    async def test_exact_13_sheet_contract_and_vector_privacy(self) -> None:
        analytics = SimpleNamespace(
            users=[],
            events=[],
            projects=[],
            department_stats=[],
            direction_stats=[],
        )
        health = SimpleNamespace(
            pulse=67,
            pulse_label="Стабильно",
            pulse_coverage=42,
            pulse_sample_size=8,
            vector_dimensions=[SimpleNamespace(label="Энергия", value=70, delta=2)],
        )
        efficiency = SimpleNamespace(score=74, label="Устойчиво")
        async with self.session_factory() as session:
            content = await build_executive_workbook(
                session,
                analytics=analytics,
                health=health,
                efficiency=efficiency,
                start_date=date(2026, 7, 21),
                end_date=date(2026, 8, 19),
                period_label="30 дней",
            )

        wb = load_workbook(BytesIO(content), data_only=True)
        self.assertEqual(wb.sheetnames, SHEET_NAMES)
        rendered = " ".join(
            str(cell.value or "")
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for cell in row
        ).casefold()
        # Executive export may explain the privacy rule in prose, but it must
        # never expose the underlying technical/raw field names or PII columns.
        for forbidden in (
            "telegram_id",
            "user_id",
            "raw_answer",
            "raw_answers",
            "personal_notes",
            "phone_number",
            "email_address",
        ):
            self.assertNotIn(forbidden, rendered)
        self.assertIn("aggregate only", rendered)
        self.assertIn("meaningful activity", rendered)
        self.assertEqual(len(wb.sheetnames), 13)


if __name__ == "__main__":
    unittest.main()
