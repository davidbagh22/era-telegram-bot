from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine, inspect

from app.config import Settings
from app.database import Base
from app.database.management_models import AdminSurvey, AdminSurveyResponse
from app.database.models import User
from app.services.scheduler_service import create_scheduler
from app.services.survey_excel_service import build_survey_workbook
from app.services.survey_service import MONTHLY_SURVEY_QUESTIONS, parse_survey_text, questions_payload


def test_admin_survey_tables_are_registered() -> None:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    tables = set(inspect(engine).get_table_names())
    assert "admin_surveys" in tables
    assert "admin_survey_responses" in tables


def test_monthly_survey_template_has_management_questions() -> None:
    assert len(MONTHLY_SURVEY_QUESTIONS) >= 6
    assert any("коммуникации" in question.lower() for question in MONTHLY_SURVEY_QUESTIONS)
    assert any("улучш" in question.lower() for question in MONTHLY_SURVEY_QUESTIONS)


def test_parse_admin_survey_text() -> None:
    title, description, questions = parse_survey_text(
        "Пульс команды\nКороткая обратная связь\n---\nЧто работает?\nЧто улучшить?"
    )
    assert title == "Пульс команды"
    assert description == "Короткая обратная связь"
    assert questions == [{"text": "Что работает?"}, {"text": "Что улучшить?"}]


def test_survey_workbook_is_readable() -> None:
    survey = AdminSurvey(
        id=1,
        title="Пульс ЭРА",
        description="Тест",
        questions_json=questions_payload(["Что работает?", "Что улучшить?"]),
        audience_type="approved",
        audience_filter_json={},
        status="sent",
        is_monthly=True,
    )
    response = AdminSurveyResponse(
        survey_id=1,
        user_id=1,
        answers_json=[{"question": "Что работает?", "answer": "Команда"}],
        status="completed",
    )
    user = User(id=1, telegram_id=1001, first_name="Анна", last_name="ЭРА")
    content = build_survey_workbook([survey], [response], [user])
    assert content.startswith(b"PK")
    workbook = load_workbook(BytesIO(content))
    assert {"Сводка", "Опросы", "Ответы", "По вопросам"}.issubset(set(workbook.sheetnames))


def test_scheduler_includes_monthly_surveys_job() -> None:
    settings = Settings(bot_token="1234567890:ABC", admin_ids=[1], timezone="Asia/Yerevan")
    scheduler = create_scheduler(object(), settings, object())
    assert "monthly-surveys" in {job.id for job in scheduler.get_jobs()}
