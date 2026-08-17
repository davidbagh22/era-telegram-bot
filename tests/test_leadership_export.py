from __future__ import annotations

import unittest
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from io import BytesIO

from app.api.deps import get_bot, get_current_user, get_session, get_settings
from app.api.v1.router import api_router
from app.config import Settings
from app.services.excel_service import build_leadership_workbook


def _analytics():
    return SimpleNamespace(
        vacancies_open=2,
        applications_by_status={"submitted": 1, "approved": 1},
        active_leaders=3,
        open_blockers=1,
        avg_blocker_resolution_hours=12.5,
        goals_active=2,
        goals_completed=1,
        goals_overdue=0,
        goal_completion_rate=100.0,
        reports_expected=3,
        reports_submitted=2,
        reporting_discipline_rate=66.7,
        leadership_health_score=83.3,
    )


class LeadershipWorkbookTests(unittest.TestCase):
    def test_build_leadership_workbook_produces_valid_xlsx(self) -> None:
        content = build_leadership_workbook(_analytics(), applications=[], attention_items=[])
        wb = load_workbook(BytesIO(content))
        self.assertIn("Сводка", wb.sheetnames)
        self.assertIn("Заявки по статусам", wb.sheetnames)
        ws = wb["Сводка"]
        self.assertEqual(ws["A1"].value, "Показатель")

    def test_build_leadership_workbook_with_rows(self) -> None:
        application = SimpleNamespace(
            id=1, office_title="Лидер Медиа", user_name="Anna A", status="submitted", submitted_at="2026-08-01"
        )
        item = SimpleNamespace(id=1, type="leader_blocker", severity="high", owner_id=2, responsible_id=None, status="open")
        content = build_leadership_workbook(_analytics(), applications=[application], attention_items=[item])
        wb = load_workbook(BytesIO(content))
        ws = wb["Заявки"]
        self.assertEqual(ws["B2"].value, "Лидер Медиа")


def _admin(**overrides) -> SimpleNamespace:
    defaults = dict(
        id=1, telegram_id=555, role="admin", is_blocked=False, is_archived=False, permission_grants=[]
    )
    defaults.update(overrides)
    return SimpleNamespace(**defaults)


def _build_app(user, session: SimpleNamespace) -> FastAPI:
    app = FastAPI()
    app.include_router(api_router)
    app.dependency_overrides[get_current_user] = lambda: user

    async def _session_override():
        yield session

    app.dependency_overrides[get_session] = _session_override
    app.dependency_overrides[get_settings] = lambda: Settings(bot_token="1234567890:test-token")
    app.dependency_overrides[get_bot] = lambda: None
    return app


class LeadershipExportApiTests(unittest.TestCase):
    def test_export_endpoint_returns_xlsx(self) -> None:
        session = SimpleNamespace()
        app = _build_app(_admin(), session)
        client = TestClient(app)
        with (
            patch("app.api.v1.admin.build_leadership_analytics", new=AsyncMock(return_value=_analytics())),
            patch(
                "app.api.v1.admin.position_management_service.list_all_applications",
                new=AsyncMock(return_value=[]),
            ),
            patch(
                "app.api.v1.admin.leadership_report_service.list_attention_items",
                new=AsyncMock(return_value=[]),
            ),
        ):
            response = client.get("/api/v1/admin/leadership/analytics/export.xlsx")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_export_forbidden_for_participant(self) -> None:
        session = SimpleNamespace()
        participant = SimpleNamespace(
            id=9, telegram_id=999, role="participant", is_blocked=False, is_archived=False, permission_grants=[]
        )
        app = _build_app(participant, session)
        client = TestClient(app)
        response = client.get("/api/v1/admin/leadership/analytics/export.xlsx")
        self.assertEqual(response.status_code, 403)


if __name__ == "__main__":
    unittest.main()
