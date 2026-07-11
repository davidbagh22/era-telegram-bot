from __future__ import annotations

import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.services.excel_service import build_analytics_workbook


ROOT = Path(__file__).resolve().parents[1]


def source(path: str) -> str:
    return (ROOT / path).read_text(encoding="utf-8")


class FullBotFlowContractTests(unittest.TestCase):
    def assertContainsAll(self, text: str, markers: list[str]) -> None:
        for marker in markers:
            self.assertIn(marker, text, msg=f"Не найден обязательный маркер: {marker}")

    def test_registration_flow_is_complete(self) -> None:
        text = source("app/handlers/registration.py")
        self.assertContainsAll(text, [
            "RegistrationStates.birth_date", "parse_birth_date", "calculate_age",
            "RegistrationStates.photo", "photo_file_id", "social_url",
        ])

    def test_participant_main_modules_are_registered(self) -> None:
        text = source("app/handlers/participant/__init__.py")
        self.assertContainsAll(text, [
            "navigation.router", "commands_ready.router", "task_block2.router",
            "projects_block5.router", "directions_block7.router",
            "events_stability_block8.router", "event_activities_block15.router",
            "auction_block17.router", "partner_offers_block16.router",
        ])

    def test_admin_main_modules_are_registered(self) -> None:
        text = source("app/handlers/admin/__init__.py")
        self.assertContainsAll(text, [
            "management_ready.router", "rights_block6.router", "task_review_block2.router",
            "projects_block5_decision.router", "event_registration_block14.router",
            "event_activities_block15.router", "auction_block17.router",
            "partner_offers_block16.router",
        ])

    def test_roles_and_permissions_flow_exists(self) -> None:
        text = source("app/handlers/admin/rights_block6.py")
        self.assertContainsAll(text, ["permission", "role", "is_blocked", "is_archived"])

    def test_self_service_directions_flow_exists(self) -> None:
        text = source("app/handlers/participant/directions_block7.py")
        self.assertContainsAll(text, ["UserDepartment", "UserDirection", "department", "direction"])

    def test_tasks_have_submission_review_and_points(self) -> None:
        participant = source("app/handlers/participant/task_block2.py")
        admin = source("app/handlers/admin/task_review_block2.py")
        self.assertContainsAll(participant, ["TaskSubmission", "task:result:", "admin:tasksub:approve:"])
        self.assertContainsAll(admin, ["approve", "reject", "add_points"])

    def test_project_to_event_flow_exists(self) -> None:
        text = source("app/handlers/participant/project_event_flow.py")
        self.assertContainsAll(text, ["project_id", "Event", "participant_limit", "points_for_visit"])

    def test_event_poster_helper_is_used(self) -> None:
        helper = source("app/services/event_card.py")
        self.assertContainsAll(helper, ["send_event_card", "poster_file_id", "answer_photo", "answer"])

    def test_event_registration_and_attendance_flow_exists(self) -> None:
        participant = source("app/handlers/participant/events_stability_block8.py")
        admin = source("app/handlers/admin/event_registration_block14.py")
        self.assertContainsAll(participant, ["EventRegistration", "participant_limit", "registration"])
        self.assertContainsAll(admin, ["ATTENDED", "NO_SHOW", "add_points"])

    def test_event_activities_proof_flow_exists(self) -> None:
        participant = source("app/handlers/participant/event_activities_block15.py")
        admin = source("app/handlers/admin/event_activities_block15.py")
        self.assertContainsAll(participant, ["EventActivitySubmission", "photo", "link", "text", "file"])
        self.assertContainsAll(admin, ["approve", "reject", "add_points"])

    def test_partner_offers_application_flow_exists(self) -> None:
        participant = source("app/handlers/participant/partner_offers_block16.py")
        admin = source("app/handlers/admin/partner_offers_block16.py")
        self.assertContainsAll(participant, ["PartnerOfferApplication", "total_points", "offer:view:"])
        self.assertContainsAll(admin, ["PartnerOfferApplication", "approve", "reject", "add_points"])

    def test_auction_has_time_leader_and_transfer(self) -> None:
        participant = source("app/handlers/participant/auction_block17.py")
        admin = source("app/handlers/admin/auction_block17.py")
        self.assertContainsAll(participant, ["Последняя ставка", "Лидер:", "Осталось:", "with_for_update", "total_points"])
        self.assertContainsAll(admin, ["ends_at", "Подтвердить победителя", "Лот передан победителю", "points=-winner_bid.amount"])

    def test_panel_and_bind_have_access_guards(self) -> None:
        panel = source("app/handlers/admin/management_ready.py")
        bind = source("app/handlers/chat_binding.py")
        self.assertContainsAll(panel, ["_guard", "_is_admin", "Role.ADMIN"])
        self.assertContainsAll(bind, ["_can_bind", "Role.ADMIN", "только администратор"])

    def test_excel_is_valid_for_empty_database(self) -> None:
        content = build_analytics_workbook([], [], [], {})
        workbook = load_workbook(BytesIO(content))
        required = {"Сводка", "Участники", "Департаменты", "Направления", "Мероприятия", "Проекты", "Цели месяца", "Организации"}
        self.assertTrue(required.issubset(set(workbook.sheetnames)))
        self.assertGreater(len(content), 1000)

    def test_excel_contains_management_analytics(self) -> None:
        text = source("app/services/excel_service.py")
        self.assertContainsAll(text, ["Воронка роста", "Рейтинг активности", "Возраст", "Город", "Медианный баланс", "Топ-10 активистов"])


if __name__ == "__main__":
    unittest.main()
