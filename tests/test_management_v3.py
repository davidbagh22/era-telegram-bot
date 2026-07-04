import inspect
import unittest
from datetime import date, datetime, time
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.database.models import MonthlyGoal, OrganizationContact
from app.handlers import chat
from app.handlers.admin import management_v3
from app.handlers.participant import profile_v3
from app.keyboards.participant import journey_keyboard, main_inline_keyboard
from app.services.excel_service import build_analytics_workbook
from app.services.resume_service import build_era_resume


class ManagementV3Tests(unittest.TestCase):
    def _user(self):
        department = SimpleNamespace(department=SimpleNamespace(name="Внутренние связи"))
        direction = SimpleNamespace(direction=SimpleNamespace(name="Лидерство"))
        return SimpleNamespace(
            id=1,
            first_name="Анна",
            last_name="Иванова",
            username="anna",
            age=21,
            city="Ереван",
            email="anna@example.com",
            phone="+37400000000",
            role="participant",
            participation_status="active_member",
            application_status="approved",
            created_at=datetime(2026, 7, 1, 12, 0),
            departments=[department],
            directions=[direction],
        )

    def test_management_models_exist(self):
        self.assertEqual(MonthlyGoal.__tablename__, "monthly_goals")
        self.assertEqual(OrganizationContact.__tablename__, "organization_contacts")

    def test_workbook_is_russian_and_well_structured(self):
        user = self._user()
        event = SimpleNamespace(
            title="Встреча",
            event_date=date(2026, 7, 20),
            event_time=time(18, 30),
            location="Ереван",
            status="completed",
            participant_limit=30,
            points_for_visit=10,
        )
        project = SimpleNamespace(
            title="Проект",
            author_id=1,
            status="in_progress",
            venue_status="approved",
            proposed_date=date(2026, 8, 1),
            proposed_time=time(12, 0),
            created_at=datetime(2026, 7, 2, 10, 0),
        )
        goal = SimpleNamespace(
            period="2026-07",
            scope_type="department",
            scope_name="Внутренние связи",
            title="Провести встречи",
            metric="Мероприятия",
            target_value=4,
            actual_value=2,
            status="active",
            notes=None,
        )
        contact = SimpleNamespace(
            organization="Партнёр",
            contact_name="Иван",
            position_primary="Директор",
            position_secondary=None,
            email="mail@example.com",
            phone="+374",
            notes=None,
        )
        content = build_analytics_workbook(
            [user],
            [event],
            [project],
            {1: 120},
            department_stats=[{
                "name": "Внутренние связи",
                "members": 1,
                "directions": 3,
                "projects": 1,
                "events": 1,
                "tasks_completed": 2,
                "goals_active": 1,
                "goals_completed": 0,
            }],
            direction_stats=[{
                "department": "Внутренние связи",
                "name": "Лидерство",
                "members": 1,
                "projects": 1,
                "events": 1,
                "tasks_completed": 2,
                "goals_active": 1,
                "goals_completed": 0,
            }],
            goals=[goal],
            contacts=[contact],
        )
        workbook = load_workbook(BytesIO(content))
        self.assertEqual(
            workbook.sheetnames,
            ["Сводка", "Участники", "Департаменты", "Направления", "Мероприятия", "Проекты", "Цели", "Организации"],
        )
        self.assertEqual(workbook["Сводка"]["A1"].value, "ЭРА · управленческая аналитика")
        self.assertEqual(workbook["Участники"]["H4"].value, "Участник")
        self.assertEqual(workbook["Участники"].freeze_panes, "A4")
        self.assertGreater(workbook["Участники"].column_dimensions["A"].width, 10)
        self.assertEqual(workbook["Департаменты"]["A4"].value, "Внутренние связи")

    def test_resume_is_pdf(self):
        user = self._user()
        pdf = build_era_resume(
            user,
            [SimpleNamespace(
                item_type="certificate",
                title="Сертификат",
                description="За участие",
                issued_at=date(2026, 7, 1),
            )],
            {"points": 100, "events": 2, "projects": 1, "tasks": 3},
        )
        self.assertTrue(pdf.startswith(b"%PDF"))

    def test_cabinet_is_compact_and_direction_callback_is_live(self):
        markup = journey_keyboard()
        callbacks = [
            button.callback_data
            for row in markup.inline_keyboard
            for button in row
            if button.callback_data
        ]
        self.assertEqual(len(callbacks), 5)
        self.assertIn("profile:departments", callbacks)
        self.assertNotIn("cabinet:direction:add", callbacks)

        main_callbacks = [
            button.callback_data
            for row in main_inline_keyboard()
            for button in row
            if button.callback_data
        ]
        self.assertIn("cabinet:rating", main_callbacks)

    def test_new_routers_own_required_actions(self):
        admin_source = inspect.getsource(management_v3)
        profile_source = inspect.getsource(profile_v3)
        chat_source = inspect.getsource(chat)
        for callback in (
            "admin:analytics",
            "admin:goals",
            "admin:contacts",
            "admin:content",
            "admin:chat_broadcast",
        ):
            self.assertIn(callback, admin_source)
        self.assertIn("department:apply:start", profile_source)
        self.assertIn("portfolio:resume", profile_source)
        self.assertIn('Command("bind")', chat_source)
        self.assertIn("manual.lstrip", inspect.getsource(__import__("app.handlers.admin.panel", fromlist=["panel"])))


if __name__ == "__main__":
    unittest.main()
