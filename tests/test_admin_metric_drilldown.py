from __future__ import annotations

import unittest
from datetime import date, datetime, time, timezone
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.api.v1.admin_drilldown import _workbook_bytes
from app.database.base import Base
from app.database.models import (
    Event,
    EventRegistration,
    PointTransaction,
    Project,
    Task,
    TaskSubmission,
    User,
)
from app.services.admin_dashboard_service import dashboard_metrics
from app.services.admin_metric_drilldown_service import build_metric_drilldown
from app.utils.constants import (
    ApplicationStatus,
    EventStatus,
    ProjectStatus,
    RegistrationStatus,
    TaskStatus,
)


class AdminMetricDrilldownTests(unittest.IsolatedAsyncioTestCase):
    async def asyncSetUp(self) -> None:
        self.engine = create_async_engine("sqlite+aiosqlite:///:memory:")
        async with self.engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)
        self.factory = async_sessionmaker(self.engine, expire_on_commit=False)

    async def asyncTearDown(self) -> None:
        await self.engine.dispose()

    async def _user(self, session, telegram_id: int, name: str, status=ApplicationStatus.APPROVED):
        user = User(
            telegram_id=telegram_id,
            first_name=name,
            application_status=status,
        )
        session.add(user)
        await session.flush()
        return user

    async def test_each_live_kpi_total_equals_exact_drilldown_rows(self) -> None:
        async with self.factory() as session:
            admin = await self._user(session, 10001, "Admin")
            active = await self._user(session, 10002, "Active")
            await self._user(session, 10003, "Pending", ApplicationStatus.PENDING)

            session.add(
                PointTransaction(
                    user_id=active.id,
                    points=20,
                    reason="verified task",
                    source_type="task_completion",
                    source_id=501,
                    idempotency_key="drilldown:active:1",
                    created_at=datetime.now(timezone.utc),
                )
            )
            project = Project(
                author_id=admin.id,
                title="Active project",
                short_description="test",
                status=ProjectStatus.IN_PROGRESS,
            )
            session.add(project)
            event = Event(
                title="Live event",
                description="test",
                event_date=date.today(),
                event_time=time(18, 0),
                location="ERA",
                format="offline",
                status=EventStatus.PUBLISHED,
                created_by=admin.id,
            )
            session.add(event)
            await session.flush()
            session.add(
                EventRegistration(
                    event_id=event.id,
                    user_id=active.id,
                    status=RegistrationStatus.REGISTERED,
                )
            )
            task = Task(
                title="Review task",
                description="test",
                creator_id=admin.id,
                deadline=datetime.now(timezone.utc),
                points=20,
                status=TaskStatus.REVIEW,
            )
            session.add(task)
            await session.flush()
            session.add(
                TaskSubmission(
                    task_id=task.id,
                    user_id=active.id,
                    text="done",
                    status="pending",
                )
            )
            await session.commit()

            dashboard = await dashboard_metrics(session)
            for metric in (
                "current_roster",
                "active_base",
                "projects_active",
                "events_live",
                "event_registrations",
                "task_results",
            ):
                detail = await build_metric_drilldown(session, metric)
                self.assertEqual(
                    detail.total,
                    dashboard.values[metric],
                    f"{metric} KPI and drill-down diverged",
                )
                self.assertEqual(detail.total, len(detail.rows))

    async def test_export_contains_exact_same_number_of_data_rows(self) -> None:
        async with self.factory() as session:
            await self._user(session, 20001, "One")
            await self._user(session, 20002, "Two")
            await session.commit()
            detail = await build_metric_drilldown(session, "current_roster")

        content = _workbook_bytes(detail.label, detail.rows)
        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb["Данные"]
        exported_rows = [
            row
            for row in ws.iter_rows(min_row=4, values_only=True)
            if any(value not in (None, "") for value in row)
        ]
        self.assertEqual(len(exported_rows), detail.total)


if __name__ == "__main__":
    unittest.main()
