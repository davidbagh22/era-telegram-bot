from __future__ import annotations

from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from app.services.excel_quality_service import finalize_business_workbook
from app.services.excel_report_service import add_health_sheets, build_development_workbook, polish_workbook


ROOT = Path(__file__).resolve().parents[1]


def _read(path: str) -> str:
    return (ROOT / path).read_text(encoding="utf-8")


def _workbook_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_pulse_is_privacy_safe_aggregate_not_person_score() -> None:
    source = _read("app/services/organization_health_service.py")
    extended = _read("app/services/organization_health_extended_service.py")
    analytics = _read("app/services/development_analytics.py")

    assert "community_analytics(session, period_days=30)" in source
    assert 'pulse = None if vector_suppressed' in source
    assert "raw ответы и личные заметки не используются" in source
    assert "traits_json" not in source
    assert "strengths_json" not in source
    assert "personal_notes" not in source
    assert "MINIMUM_COHORT = 5" in analytics
    assert "build_organization_health(session)" in extended
    assert "не влияют на Пульс" in extended
    assert "traits_json" not in extended
    assert "personal_notes" not in extended


def test_admin_health_endpoints_and_ui_are_exposed() -> None:
    api = _read("app/api/v1/admin_analytics_details.py")
    dashboard = _read("frontend/src/screens/admin/AdminDashboardScreen.tsx")
    overview = _read("frontend/src/screens/admin/AdminOverviewScreen.tsx")

    assert '@router.get("/health"' in api
    assert '@router.get("/health-report.xlsx")' in api
    assert '@router.get("/full-report.xlsx")' in api
    assert "build_extended_organization_health" in api
    assert "Пульс организации" in dashboard
    assert "Здоровье организации · XLSX" in dashboard
    assert "Показать все" in dashboard
    assert "<AdminDashboardScreen />" in overview
    assert "<SystemPanel />" in overview
    assert "Техническое состояние" in overview


def test_extended_health_uses_real_growth_opportunity_and_career_models() -> None:
    source = _read("app/services/organization_health_extended_service.py")

    for marker in [
        "ReferralRelationship",
        "registration_rewarded_at",
        "first_event_rewarded_at",
        "PartnerInitiative",
        "SavedOpportunity",
        "PartnerOfferApplication",
        "CareerProfile",
        "CareerPortfolioItem",
        "RecommendationRequest",
        '"referral_event_conversion"',
        '"active_opportunities"',
        '"career_profile_adoption"',
        '"portfolio_pending"',
    ]:
        assert marker in source


def test_admin_overview_replaces_separate_maintenance_and_analytics_hubs() -> None:
    admin = _read("frontend/src/screens/AdminScreen.tsx")
    overview = _read("frontend/src/screens/admin/AdminOverviewScreen.tsx")
    nav = _read("frontend/src/components/AdminBottomNav.tsx")

    assert "AdminMaintenanceScreen" not in admin
    assert "ControlSection" not in admin
    assert "CONTROL_SECTIONS" not in admin
    assert 'label: "Контроль"' not in nav
    assert 'label: "Аналитика"' not in nav
    assert 'label: "Обслуживание"' not in nav
    assert "<AdminDashboardScreen />" in overview
    assert "<SystemPanel />" in overview
    assert 'id="admin-analytics"' in overview
    for destination in [
        "participants",
        "applications",
        "development",
        "career",
        "offices",
        "projects",
        "events",
        "tasks",
        "offers",
        "data-rights",
        "surveys",
        "tools",
    ]:
        assert f'"{destination}"' in admin


def test_polish_workbook_removes_internal_ids_and_keeps_business_columns() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Мероприятия"
    ws.append(["Название", "Проект ID", "Ответственный ID", "Автор ID", "Статус"])
    ws.append(["Форум ЭРА", 7, 12, 44, "active"])

    polished = load_workbook(BytesIO(polish_workbook(_workbook_bytes(wb))))
    result = polished["Мероприятия"]
    headers = [result.cell(1, col).value for col in range(1, result.max_column + 1)]

    assert headers == ["Название", "Статус"]
    assert result.sheet_view.showGridLines is False
    assert result.freeze_panes == "A2"


def test_excel_quality_moves_filter_off_title_row() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    ws.merge_cells("A1:D1")
    ws["A1"] = "ЗДОРОВЬЕ ЭРА"
    ws.append(["Подзаголовок"])
    ws.append([])
    ws.append(["Показатель", "Значение", "Пояснение", "Оценка"])
    ws.append(["Удержание", "70%", "30 дней", 70])
    ws.auto_filter.ref = "A1:D5"

    result = load_workbook(BytesIO(finalize_business_workbook(_workbook_bytes(wb))))["Отчёт"]

    assert result.auto_filter.ref == "A4:D5"
    assert result.freeze_panes == "A5"
    assert result.sheet_view.showGridLines is False


def test_health_report_starts_with_executive_pulse_sheets() -> None:
    wb = Workbook()
    wb.active.title = "Сводка"
    wb.active.append(["Показатель", "Значение"])
    wb.active.append(["Участники", 218])

    health = SimpleNamespace(
        pulse=67,
        pulse_label="Состояние устойчивое",
        pulse_coverage=80,
        pulse_sample_size=8,
        pulse_suppressed=False,
        vector_dimensions=[
            SimpleNamespace(key="energy", label="Энергия", value=71, delta=2),
            SimpleNamespace(key="agency", label="Опора", value=65, delta=-1),
            SimpleNamespace(key="autonomy", label="Самостоятельность", value=69, delta=0),
            SimpleNamespace(key="connection", label="Связь", value=73, delta=4),
            SimpleNamespace(key="direction", label="Направление", value=58, delta=-3),
        ],
        metrics=[
            SimpleNamespace(
                key="active_30d",
                category="Вовлечённость",
                label="Активны за 30 дней",
                value=55.0,
                display="120 · 55.0%",
                note="основной показатель текущей вовлечённости",
                score=55,
            ),
            SimpleNamespace(
                key="overdue_tasks",
                category="Исполнение",
                label="Просроченных заданий",
                value=3.0,
                display="3",
                note="срок прошёл, задача не закрыта",
                score=None,
            ),
            SimpleNamespace(
                key="referral_event_conversion",
                category="Рост",
                label="Приглашённый → первое мероприятие",
                value=60.0,
                display="60.0%",
                note="реферальная конверсия",
                score=60,
            ),
        ],
        risks=["Просрочено заданий: 3."],
        period_label="операционные показатели: последние 7/30/90 дней",
        data_note="Пульс — безопасный агрегат текущего состояния.",
    )
    efficiency = SimpleNamespace(score=74, label="Система работает устойчиво")

    content = finalize_business_workbook(add_health_sheets(_workbook_bytes(wb), health, efficiency))
    result = load_workbook(BytesIO(content))

    assert result.sheetnames[:3] == ["Здоровье организации", "Пульс Мой вектор", "Все показатели"]
    assert result["Здоровье организации"]["B4"].value == 74
    assert result["Здоровье организации"]["B5"].value == 67
    assert result["Пульс Мой вектор"]["B4"].value == 67
    assert result["Все показатели"].max_row >= 7
    assert result["Все показатели"].auto_filter.ref.startswith("A4:E")


def test_suppressed_vector_export_contains_no_individual_data() -> None:
    content = finalize_business_workbook(
        build_development_workbook(
            {
                "coverage_percent": 20,
                "sample_size": 2,
                "eligible_profiles": 10,
                "minimum_cohort": 5,
                "suppressed": True,
            }
        )
    )
    wb = load_workbook(BytesIO(content))
    values = " ".join(
        str(cell.value or "")
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
    ).lower()

    assert "недостаточно" in values
    assert "личные заметки" in values
    assert "user_id" not in values
    assert "raw" not in values
