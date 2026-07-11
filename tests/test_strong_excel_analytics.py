from __future__ import annotations

import unittest
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.excel_service import build_analytics_workbook


class StrongExcelAnalyticsTests(unittest.TestCase):
    def test_empty_workbook_is_valid_and_manager_ready(self) -> None:
        content = build_analytics_workbook([], [], [], {})
        workbook = load_workbook(BytesIO(content))
        self.assertIn("Сводка", workbook.sheetnames)
        self.assertIn("Воронка роста", workbook.sheetnames)
        self.assertIn("Срезы аудитории", workbook.sheetnames)
        self.assertIn("Активность участников", workbook.sheetnames)
        self.assertEqual(workbook["Участники"]["A2"].value, "Данных пока нет")

    def test_dashboard_contains_management_metrics(self) -> None:
        user = SimpleNamespace(
            id=1,
            first_name="Давид",
            last_name="Багдасарян",
            username="davidbagh22",
            telegram_id=100,
            birth_date=None,
            age=24,
            city="Ереван",
            email="david@example.com",
            role="participant",
            participation_status="active",
            application_status="approved",
            created_at=None,
            departments=[1],
            directions=[1],
        )
        content = build_analytics_workbook([user], [], [], {1: 250})
        workbook = load_workbook(BytesIO(content), data_only=True)
        values = [workbook["Сводка"].cell(row, 1).value for row in range(1, workbook["Сводка"].max_row + 1)]
        self.assertIn("Участники с баллами", values)
        self.assertIn("Медианный баланс", values)
        self.assertEqual(workbook["Активность участников"]["D2"].value, 250)


if __name__ == "__main__":
    unittest.main()
